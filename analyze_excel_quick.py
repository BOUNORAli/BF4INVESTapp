import sys
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

file_path = r'C:\Users\PC\Downloads\test gestion boubker.xlsx'

print("=" * 100)
print("ANALYSE DU FICHIER EXCEL - GESTION BOUbKER")
print("=" * 100)

wb = load_workbook(file_path, data_only=False)

print(f"\n📊 FEUILLES: {wb.sheetnames}")
print(f"📋 Nombre: {len(wb.sheetnames)}\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print("=" * 100)
    print(f"📄 FEUILLE: {sheet_name}")
    print("=" * 100)
    print(f"Dimensions: {sheet.max_row} lignes x {sheet.max_column} colonnes\n")
    
    # En-têtes (lignes 1-5)
    print("📌 EN-TÊTES (lignes 1-5):")
    print("-" * 100)
    for row_idx in range(1, min(6, sheet.max_row + 1)):
        row = sheet[row_idx]
        values = [f"{cell.column_letter}:{str(cell.value)[:40]}" for cell in row[:15] if cell.value is not None]
        if values:
            print(f"L{row_idx}: {' | '.join(values)}")
    
    # Échantillon de données (lignes 6-20)
    print("\n📝 ÉCHANTILLON DONNÉES (lignes 6-20):")
    print("-" * 100)
    for row_idx in range(6, min(21, sheet.max_row + 1)):
        row = sheet[row_idx]
        values = []
        for cell in row[:12]:  # 12 premières colonnes
            if cell.value is not None:
                val = str(cell.value)
                if len(val) > 25:
                    val = val[:25] + "..."
                if cell.data_type == 'f':
                    val = f"[FORMULE: {val[:30]}]"
                values.append(f"{cell.column_letter}:{val}")
        if values:
            print(f"L{row_idx}: {' | '.join(values)}")
    
    # Formules uniques (échantillon)
    print("\n🔢 FORMULES UNIQUES (échantillon 15 premières):")
    print("-" * 100)
    formulas_seen = set()
    count = 0
    for row in sheet.iter_rows(min_row=1, max_row=min(200, sheet.max_row)):
        for cell in row:
            if cell.data_type == 'f' and cell.value not in formulas_seen:
                formulas_seen.add(cell.value)
                print(f"{cell.coordinate}: {cell.value}")
                count += 1
                if count >= 15:
                    break
        if count >= 15:
            break
    
    # Colonnes avec leurs types
    print("\n📊 COLONNES ET TYPES (analyse des 50 premières lignes):")
    print("-" * 100)
    col_info = {}
    for row in sheet.iter_rows(min_row=1, max_row=min(50, sheet.max_row)):
        for cell in row:
            if cell.value is not None:
                col = cell.column_letter
                if col not in col_info:
                    col_info[col] = {'types': set(), 'formulas': 0, 'values': 0}
                if cell.data_type == 'f':
                    col_info[col]['formulas'] += 1
                else:
                    col_info[col]['values'] += 1
                col_info[col]['types'].add(type(cell.value).__name__)
    
    for col in sorted(col_info.keys())[:20]:  # 20 premières colonnes
        info = col_info[col]
        print(f"Col {col}: Types={list(info['types'])[:3]}, Formules={info['formulas']}, Valeurs={info['values']}")

wb.close()
print("\n" + "=" * 100)
print("ANALYSE TERMINÉE")
print("=" * 100)





