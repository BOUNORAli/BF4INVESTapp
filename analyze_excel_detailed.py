import sys
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

file_path = r'C:\Users\PC\Downloads\test gestion boubker.xlsx'

print("=" * 100)
print("ANALYSE DÉTAILLÉE DU FICHIER EXCEL - GESTION BOUbKER")
print("=" * 100)

wb = load_workbook(file_path, data_only=False)

print(f"\n📊 NOMBRE DE FEUILLES: {len(wb.sheetnames)}")
print(f"📋 NOMS DES FEUILLES: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print("=" * 100)
    print(f"📄 FEUILLE: {sheet_name}")
    print("=" * 100)
    print(f"Dimensions: {sheet.max_row} lignes x {sheet.max_column} colonnes\n")
    
    # Analyser les en-têtes (premières lignes)
    print("📌 EN-TÊTES ET STRUCTURE:")
    print("-" * 100)
    for row_idx in range(1, min(10, sheet.max_row + 1)):
        row = sheet[row_idx]
        row_data = []
        for cell in row:
            if cell.value is not None:
                cell_info = f"{cell.coordinate}:{str(cell.value)[:50]}"
                if cell.data_type == 'f':
                    cell_info += " [FORMULE]"
                row_data.append(cell_info)
        if row_data:
            print(f"Ligne {row_idx}: {' | '.join(row_data[:8])}")
    
    # Analyser les formules importantes
    print("\n🔢 FORMULES IMPORTANTES (échantillon):")
    print("-" * 100)
    formula_count = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula
                formula_count += 1
                if formula_count <= 20:  # Afficher les 20 premières
                    print(f"{cell.coordinate}: {cell.value}")
    
    print(f"\nTotal formules trouvées: {formula_count}")
    
    # Analyser les colonnes et leur contenu
    print("\n📊 ANALYSE DES COLONNES (premières lignes de données):")
    print("-" * 100)
    
    # Trouver la ligne d'en-tête (généralement ligne 1-4)
    header_row = None
    for row_idx in range(1, min(6, sheet.max_row + 1)):
        row = sheet[row_idx]
        non_empty = sum(1 for cell in row if cell.value is not None)
        if non_empty > 5:  # Si plus de 5 cellules remplies, probablement un en-tête
            header_row = row_idx
            break
    
    if header_row:
        print(f"Ligne d'en-tête probable: {header_row}")
        headers = []
        for cell in sheet[header_row]:
            headers.append(cell.value if cell.value else f"Col{cell.column_letter}")
        print(f"Colonnes: {headers[:15]}...")  # Afficher les 15 premières
    
    # Analyser quelques lignes de données
    print("\n📝 ÉCHANTILLON DE DONNÉES (lignes 5-15):")
    print("-" * 100)
    for row_idx in range(5, min(16, sheet.max_row + 1)):
        row = sheet[row_idx]
        row_values = []
        for cell in row[:10]:  # Premières 10 colonnes
            if cell.value is not None:
                val = str(cell.value)
                if len(val) > 30:
                    val = val[:30] + "..."
                row_values.append(f"{cell.column_letter}:{val}")
        if row_values:
            print(f"Ligne {row_idx}: {' | '.join(row_values)}")
    
    # Analyser les références de cellules dans les formules
    print("\n🔗 RÉFÉRENCES DE CELLULES DANS LES FORMULES:")
    print("-" * 100)
    referenced_cells = set()
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula = str(cell.value)
                # Extraire les références de cellules (format $A$1, A1, etc.)
                import re
                refs = re.findall(r'\$?[A-Z]+\$?\d+', formula)
                referenced_cells.update(refs)
                if len(referenced_cells) > 50:
                    break
        if len(referenced_cells) > 50:
            break
    
    print(f"Cellules référencées (échantillon): {list(referenced_cells)[:20]}")
    
    # Analyser les types de données
    print("\n📈 TYPES DE DONNÉES:")
    print("-" * 100)
    data_types = {}
    for row in sheet.iter_rows(min_row=5, max_row=min(100, sheet.max_row)):
        for cell in row:
            if cell.value is not None:
                dtype = type(cell.value).__name__
                data_types[dtype] = data_types.get(dtype, 0) + 1
    
    for dtype, count in sorted(data_types.items(), key=lambda x: x[1], reverse=True):
        print(f"{dtype}: {count} occurrences")

wb.close()
print("\n" + "=" * 100)
print("ANALYSE TERMINÉE")
print("=" * 100)



