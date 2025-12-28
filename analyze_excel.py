import sys
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Installation de openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import load_workbook

file_path = r'C:\Users\PC\Downloads\test gestion boubker.xlsx'

print("=" * 80)
print("ANALYSE DÉTAILLÉE DU FICHIER EXCEL")
print("=" * 80)

wb = load_workbook(file_path, data_only=False)

print(f"\n📊 NOMBRE DE FEUILLES: {len(wb.sheetnames)}")
print(f"📋 NOMS DES FEUILLES: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print("\n" + "=" * 80)
    print(f"📄 FEUILLE: {sheet_name}")
    print("=" * 80)
    print(f"Dimensions: {sheet.max_row} lignes x {sheet.max_column} colonnes")
    
    # Analyser toutes les cellules avec contenu
    print("\n📝 CONTENU DÉTAILLÉ:")
    print("-" * 80)
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                cell_type = "FORMULE" if isinstance(cell.value, str) and cell.value.startswith('=') else "VALEUR"
                if cell_type == "FORMULE":
                    print(f"{cell.coordinate}: {cell.value}")
                else:
                    print(f"{cell.coordinate}: {cell.value} ({type(cell.value).__name__})")
    
    # Analyser les formules séparément
    print("\n🔢 FORMULES DÉTECTÉES:")
    print("-" * 80)
    formulas_found = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula
                formulas_found = True
                print(f"{cell.coordinate}: {cell.value}")
    if not formulas_found:
        print("Aucune formule détectée")
    
    # Analyser les styles et formats
    print("\n🎨 STYLES ET FORMATS:")
    print("-" * 80)
    for row in sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row)):
        for cell in row:
            if cell.value is not None:
                print(f"{cell.coordinate}: format={cell.number_format}, alignement={cell.alignment}")

wb.close()




